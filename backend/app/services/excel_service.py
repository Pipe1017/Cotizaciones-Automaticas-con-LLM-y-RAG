"""
Genera el Excel de cotización OPEX desde código (sin template).
Dinámico: soporta cualquier número de ítems.
Paleta: Navy #0F2560, Orange #F97316, Light #F1F5F9.
"""

from io import BytesIO
from datetime import date
from decimal import Decimal
from copy import copy
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta OPEX ────────────────────────────────────────────────
NAVY   = "0F2560"
ORANGE = "F97316"
WHITE  = "FFFFFF"
LIGHT  = "F1F5F9"
GRAY   = "E2E8F0"
TEXT   = "1E293B"

# ── Columnas de la tabla de ítems ──────────────────────────────
# A=1 margen | B=2 # | C=3 referencia | D=4 descripción |
# E=5 SAP | F=6 marca | G=7 qty | H=8 precio unit | I=9 precio total
COL_WIDTHS = {
    1: 2,    # A — margen izquierdo
    2: 6,    # B — ITEM #
    3: 14,   # C — REFERENCIA
    4: 34,   # D — DESCRIPCIÓN
    5: 16,   # E — COD. PROVEEDOR
    6: 14,   # F — MARCA
    7: 8,    # G — QTD
    8: 17,   # H — PRECIO UNIT. USD
    9: 17,   # I — PRECIO TOTAL USD
    10: 2,   # J — margen derecho
}
LAST_COL = 9
FIRST_DATA_COL = 2


# ── Helpers de estilo ──────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=TEXT, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color=ORANGE, style="medium"):
    return Border(bottom=Side(style=style, color=color))

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _money(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = '#,##0.00'
    return c

def _merge_row(ws, row, col_start, col_end):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    return ws.cell(row=row, column=col_start)


def _section_header(ws, row, text):
    """Fila de sección: fondo navy, texto blanco, toda la fila."""
    _merge_row(ws, row, FIRST_DATA_COL, LAST_COL)
    c = ws.cell(row=row, column=FIRST_DATA_COL, value=text)
    c.font      = _font(bold=True, size=9, color=WHITE)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 16


def _apply_item_row_style(ws, row, zebra=False):
    """Aplica estilo de fila de datos (alternada)."""
    bg = LIGHT if zebra else WHITE
    thin = _border()
    for col in range(FIRST_DATA_COL, LAST_COL + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = _fill(bg)
        c.border    = thin
        c.font      = _font(size=9)
        c.alignment = _align("center") if col in (2, 7, 8, 9) else _align("left", wrap=True)
    ws.row_dimensions[row].height = 30


# ── Generador principal ────────────────────────────────────────

def generate_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización OPEX"

    # Anchos de columna
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Configuración de impresión
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_margins.left = ws.page_margins.right = 0.5

    row = 1

    # ── HEADER ────────────────────────────────────────────────────
    # Barra navy superior
    ws.row_dimensions[row].height = 6
    for col in range(1, LAST_COL + 2):
        ws.cell(row=row, column=col).fill = _fill(NAVY)
    row += 1

    # Fila con logos y título
    ws.row_dimensions[row].height = 52
    # [B] Área logo OPEX
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    logo_cell = ws.cell(row=row, column=2, value="[ Logo OPEX ]")
    logo_cell.font      = _font(bold=True, size=9, color=NAVY)
    logo_cell.alignment = _align("center", "center")
    logo_cell.fill      = _fill("F8FAFC")

    # [D-F] Título central
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    title = ws.cell(row=row, column=4, value="COTIZACIÓN / OFERTA OPEX SAS")
    title.font      = _font(bold=True, size=14, color=NAVY)
    title.alignment = _align("center", "center")
    title.fill      = _fill("F8FAFC")

    # [H-I] Área logo cliente
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    client_logo = ws.cell(row=row, column=8, value="[ Logo Cliente ]")
    client_logo.font      = _font(size=9, color="94A3B8")
    client_logo.alignment = _align("center", "center")
    client_logo.fill      = _fill("F8FAFC")

    # Borde inferior naranja del header
    for col in range(FIRST_DATA_COL, LAST_COL + 1):
        ws.cell(row=row, column=col).border = _border_bottom(ORANGE, "thick")
    row += 1

    # Línea separadora
    ws.row_dimensions[row].height = 4
    for col in range(1, LAST_COL + 2):
        ws.cell(row=row, column=col).fill = _fill(ORANGE)
    row += 1

    # ── INFO COTIZACIÓN ───────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    # Label + valor en pares
    pairs = [
        (2, "Nº Cotización:", 3, data.get("numero_cotizacion", "")),
        (5, "Fecha:", 6, data.get("fecha", date.today()).strftime("%d/%m/%Y") if hasattr(data.get("fecha"), "strftime") else ""),
        (8, "Páginas:", 9, "1"),
    ]
    for lcol, ltext, vcol, vval in pairs:
        lc = ws.cell(row=row, column=lcol, value=ltext)
        lc.font = _font(bold=True, size=9, color=NAVY)
        lc.alignment = _align("right")
        vc = ws.cell(row=row, column=vcol, value=vval)
        vc.font = _font(size=9)
        vc.alignment = _align("left")
        vc.border = Border(bottom=Side(style="thin", color=GRAY))
    row += 1

    ws.row_dimensions[row].height = 20
    pairs2 = [
        (2, "Cliente:", 3, data.get("cliente", "")),
        (6, "Contacto:", 7, data.get("contacto_nombre", "")),
    ]
    for lcol, ltext, vcol, vval in pairs2:
        lc = ws.cell(row=row, column=lcol, value=ltext)
        lc.font = _font(bold=True, size=9, color=NAVY)
        lc.alignment = _align("right")
        ws.merge_cells(start_row=row, start_column=vcol, end_row=row, end_column=vcol + 1)
        vc = ws.cell(row=row, column=vcol, value=vval)
        vc.font = _font(bold=True, size=10, color=NAVY)
        vc.alignment = _align("left")
        vc.border = Border(bottom=Side(style="medium", color=NAVY))
    row += 1

    ws.row_dimensions[row].height = 8  # espacio
    row += 1

    # ── TABLA DE ÍTEMS ────────────────────────────────────────────
    # Header de tabla
    ws.row_dimensions[row].height = 22
    headers = [
        (2, "#"),
        (3, "REFERENCIA"),
        (4, "DESCRIPCIÓN"),
        (5, "COD. PROVEEDOR"),
        (6, "MARCA"),
        (7, "QTD"),
        (8, "PRECIO UNIT. USD"),
        (9, "PRECIO TOTAL USD"),
    ]
    for col, txt in headers:
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _align("center")
        c.border    = Border(
            left=Side(style="thin", color="1E3A6E"),
            right=Side(style="thin", color="1E3A6E"),
            bottom=Side(style="medium", color=ORANGE),
        )
    row += 1

    # Ítems dinámicos — primero los normales, luego los opcionales
    items = data.get("items", [])
    normales  = [i for i in items if not i.get("opcional", False)]
    opcionales = [i for i in items if i.get("opcional", False)]

    for idx, item in enumerate(normales):
        _apply_item_row_style(ws, row, zebra=(idx % 2 == 1))
        ws.cell(row=row, column=2, value=idx + 1).font = _font(bold=True, size=9, color=NAVY)
        ws.cell(row=row, column=3, value=item.get("referencia_usa", ""))
        ws.cell(row=row, column=4, value=item.get("descripcion", ""))
        ws.cell(row=row, column=5, value=item.get("referencia_cod_proveedor", ""))
        ws.cell(row=row, column=6, value=item.get("marca") or "")
        ws.cell(row=row, column=7, value=float(item.get("cantidad", 0)))
        _money(ws, row, 8, float(item.get("precio_unitario_usd", 0)))
        _money(ws, row, 9, float(item.get("precio_total_usd", 0)))
        for col in (7, 8, 9):
            ws.cell(row=row, column=col).alignment = _align("right")
        row += 1

    # Línea de cierre de ítems normales
    if normales:
        for col in range(FIRST_DATA_COL, LAST_COL + 1):
            c = ws.cell(row=row - 1, column=col)
            c.border = Border(
                left=c.border.left, right=c.border.right,
                top=c.border.top,
                bottom=Side(style="medium", color=NAVY),
            )

    # Sección OPCIONALES
    if opcionales:
        ws.row_dimensions[row].height = 6
        row += 1
        _section_header(ws, row, "ÍTEMS OPCIONALES — No incluidos en el total")
        row += 1

        GRAY_OPT = "E2E8F0"
        for idx, item in enumerate(opcionales):
            ws.row_dimensions[row].height = 28
            for col in range(FIRST_DATA_COL, LAST_COL + 1):
                c = ws.cell(row=row, column=col)
                c.fill   = _fill(GRAY_OPT)
                c.border = _border(color="CBD5E1")
                c.font   = _font(size=9, color="64748B")
                c.alignment = _align("center") if col in (2, 7, 8, 9) else _align("left", wrap=True)
            ws.cell(row=row, column=2, value="OPC").font = _font(bold=True, size=8, color="94A3B8")
            ws.cell(row=row, column=3, value=item.get("referencia_usa", ""))
            nota = item.get("notas") or ""
            desc = item.get("descripcion", "")
            ws.cell(row=row, column=4, value=f"{desc}  [{nota}]" if nota else desc)
            ws.cell(row=row, column=5, value=item.get("referencia_cod_proveedor", ""))
            ws.cell(row=row, column=6, value=item.get("marca") or "")
            ws.cell(row=row, column=7, value="OPCIONAL")
            ws.cell(row=row, column=7).font = _font(bold=True, size=8, color="94A3B8")
            ref_c = ws.cell(row=row, column=8, value=float(item.get("precio_unitario_usd", 0)))
            ref_c.number_format = '#,##0.00'
            ref_c.alignment = _align("right")
            ref_c.font = _font(size=9, color="94A3B8")
            ws.cell(row=row, column=9, value="—").alignment = _align("center")
            row += 1

    ws.row_dimensions[row].height = 8
    row += 1

    # ── SERVICIOS DE INGENIERÍA ────────────────────────────────────
    servicios = data.get("servicios", [])
    if servicios:
        _section_header(ws, row, "SERVICIOS DE INGENIERÍA")
        row += 1
        # sub-header
        ws.row_dimensions[row].height = 16
        for col, txt in [(2, "ROL"), (4, "DESCRIPCIÓN / MOTIVO"), (7, "HORAS"), (8, "TARIFA/H"), (9, "SUBTOTAL")]:
            c = ws.cell(row=row, column=col, value=txt)
            c.font = _font(bold=True, size=8, color="64748B")
            c.fill = _fill(LIGHT)
            c.alignment = _align("center" if col in (7, 8, 9) else "left")
        row += 1
        for svc in servicios:
            ws.row_dimensions[row].height = 22
            for col in range(FIRST_DATA_COL, LAST_COL + 1):
                c = ws.cell(row=row, column=col)
                c.fill = _fill("F8FAFC")
                c.border = _border(color="E2E8F0")
                c.font = _font(size=9)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row=row, column=2, value=svc.get("nombre", "")).font = _font(bold=True, size=9, color=NAVY)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            ws.cell(row=row, column=4, value=svc.get("motivo", "")).font = _font(size=8, color="64748B")
            ws.cell(row=row, column=4).alignment = _align("left", wrap=True)
            _money(ws, row, 7, float(svc.get("horas", 0))).alignment = _align("right")
            _money(ws, row, 8, float(svc.get("tarifa_hora_usd", 0))).alignment = _align("right")
            _money(ws, row, 9, float(svc.get("subtotal_usd", 0))).alignment = _align("right")
            row += 1
        # subtotal servicios
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        lc = ws.cell(row=row, column=2, value="Subtotal Servicios")
        lc.font = _font(bold=True, size=9, color=NAVY)
        lc.fill = _fill(LIGHT)
        lc.alignment = _align("right")
        vc = ws.cell(row=row, column=9, value=float(data.get("servicios_subtotal_usd", 0)))
        vc.number_format = '#,##0.00'
        vc.font = _font(bold=True, size=9, color=NAVY)
        vc.fill = _fill(LIGHT)
        vc.alignment = _align("right")
        row += 1
        ws.row_dimensions[row].height = 6
        row += 1

    # ── TOTALES ────────────────────────────────────────────────────
    subtotal = float(data.get("subtotal_usd", 0))
    svc_sub  = float(data.get("servicios_subtotal_usd", 0))
    iva_pct  = float(data.get("iva_pct", 19))
    iva      = (subtotal + svc_sub) * iva_pct / 100
    total    = float(data.get("total_usd", 0))

    rows_totales = [("SUBTOTAL PRODUCTOS", subtotal, False)]
    if svc_sub > 0:
        rows_totales.append(("SUBTOTAL SERVICIOS", svc_sub, False))
    rows_totales += [(f"IVA {int(iva_pct)}%", iva, False), ("TOTAL USD", total, True)]

    for label, value, is_total in rows_totales:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = _font(bold=True, size=9, color=WHITE if is_total else NAVY)
        lc.fill      = _fill(NAVY if is_total else LIGHT)
        lc.alignment = _align("right")
        lc.border    = _border(color="CBD5E1")

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        vc = ws.cell(row=row, column=8, value=value)
        vc.number_format = '#,##0.00'
        vc.font      = _font(bold=True, size=11 if is_total else 9,
                             color=WHITE if is_total else TEXT)
        vc.fill      = _fill(NAVY if is_total else LIGHT)
        vc.alignment = _align("right")
        vc.border    = _border(color="CBD5E1")
        row += 1

    row += 1  # espacio

    # ── OBSERVACIONES ─────────────────────────────────────────────
    obs = data.get("observaciones", "").strip()
    if obs:
        _section_header(ws, row, "OBSERVACIONES")
        row += 1
        ws.row_dimensions[row].height = max(30, len(obs) // 3)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        oc = ws.cell(row=row, column=2, value=obs)
        oc.font      = _font(size=9)
        oc.alignment = _align("left", wrap=True)
        oc.fill      = _fill("FFFBF0")
        oc.border    = _border(color="FED7AA")
        row += 2

    # ── CONDICIONES COMERCIALES ────────────────────────────────────
    _section_header(ws, row, "CONDICIONES COMERCIALES")
    row += 1

    conditions = [
        ("Fecha de entrega",       data.get("fecha_entrega", "")),
        ("Condiciones de entrega", data.get("condiciones_entrega", "")),
        ("Condiciones de pago",    data.get("condiciones_pago", "")),
        ("Condiciones de garantía",data.get("condiciones_garantia", "")),
        ("Validez de la oferta",   data.get("validez_oferta", "30 días")),
    ]
    for label, value in conditions:
        if not value:
            continue
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = _font(bold=True, size=9, color=NAVY)
        lc.fill      = _fill(LIGHT)
        lc.alignment = _align("left")
        lc.border    = Border(
            left=Side(style="thin", color=GRAY),
            bottom=Side(style="thin", color=GRAY),
        )

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        vc = ws.cell(row=row, column=4, value=value)
        vc.font      = _font(size=9)
        vc.alignment = _align("left")
        vc.border    = Border(
            right=Side(style="thin", color=GRAY),
            bottom=Side(style="thin", color=GRAY),
        )
        row += 1

    row += 1

    # ── FIRMA ──────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    asesor  = data.get("asesor", "OPEX SAS")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    sc = ws.cell(row=row, column=2, value=asesor)
    sc.font      = _font(bold=True, size=10, color=NAVY)
    sc.alignment = _align("center")
    sc.border    = Border(top=Side(style="medium", color=NAVY))

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    ec = ws.cell(row=row, column=6, value="OPEX SAS")
    ec.font      = _font(bold=True, size=10, color=NAVY)
    ec.alignment = _align("center")
    ec.border    = Border(top=Side(style="medium", color=NAVY))
    row += 1

    # Barra naranja inferior
    row += 1
    ws.row_dimensions[row].height = 5
    for col in range(1, LAST_COL + 2):
        ws.cell(row=row, column=col).fill = _fill(ORANGE)

    # ── GUARDAR ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# Mantener compatibilidad con el código existente que llama a fill_template
def fill_template(_template_bytes: bytes, data: dict) -> bytes:
    """Wrapper: genera el Excel desde código ignorando el template."""
    return generate_excel(data)
